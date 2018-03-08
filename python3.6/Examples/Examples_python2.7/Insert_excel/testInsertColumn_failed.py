from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet import *
from openpyxl.cell import  Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
import re
import copy


def insert_rows(self, row_idx, cnt, above=False, copy_style=True, 
fill_formulae=True):
  RE_CELL  = re.compile("(?P<col>[A-Z]+)(?P<row>\d+)")
  RE_RANGE = re.compile("(?P<s_col>[A-Z]+)(?P<s_row>\d+):(?P<e_col>[A-Z]+)(?P<e_row>\d+)")
 
  row_idx = row_idx - 1 if above else row_idx
 
  # First, we shift all cells down cnt rows...
  old_cells = set()
  old_fas   = set()
  new_cells = dict()
  new_fas   = dict()
  for c in self._cells.values():
    if c.row > row_idx:
      old_coor = c.coordinate
      old_cells.add((c.row,c.col_idx))
      c.row += cnt
      new_cells[(c.row,c.col_idx)] = c
      if c.data_type == Cell.TYPE_FORMULA:
        #print("Updating formula in cell %s%d to match %s%d"%(c.column,c.row-cnt,c.column,c.row))
        c.value = re.sub(
          "(\$?[A-Z]{1,3})\$?%d"%(c.row-cnt),
          lambda m: m.group(1) + str(c.row),
          c.value
        )
        # Here, we need to properly update the formula references to reflect new row indices
        if old_coor in self.formula_attributes:
          old_fas.add(old_coor)
          fa = self.formula_attributes[old_coor].copy()
          if 'ref' in fa:
            # Handle case where cell references itself!
            if fa['ref'] == old_coor:
              fa['ref'] = c.coordinate
            # Otherwise, we need to shift the range reference down by cnt
            else:
              g = RE_RANGE.search(fa['ref']).groupdict()
              fa['ref'] = g['s_col'] + str(int(g['s_row'])+cnt) + ":" + g['e_col'] + str(int(g['e_row'])+cnt)
          new_fas[c.coordinate] = fa
 
  for coor in old_cells:
    del self._cells[coor]
  self._cells.update(new_cells)
 
  for fa in old_fas:
    del self.formula_attributes[fa]
  self.formula_attributes.update(new_fas)
 
  # Next, we need to shift all the Row Dimensions below out new rows down by cnt...
  for row in range(len(self.row_dimensions)-1+cnt,row_idx+cnt,-1):
    new_rd = copy.copy(self.row_dimensions[row-cnt])
    new_rd.index = row
    self.row_dimensions[row] = new_rd
    del self.row_dimensions[row-cnt]
 
  # Now, create our new rows, with all the pretty cells
  row_idx += 1
  for row in range(row_idx,row_idx+cnt):
    # Create a Row Dimension for our new row
    new_rd = copy.copy(self.row_dimensions[row-1])
    new_rd.index = row
    self.row_dimensions[row] = new_rd
    for col in range(1,ws.max_column):
      col = get_column_letter(col)
      cell = self.cell('%s%d'%(col,row))
      cell.value = None
      source = self.cell('%s%d'%(col,row-1))
      if copy_style:
        cell.number_format = source.number_format
        cell.font      = source.font.copy()
        cell.alignment = source.alignment.copy()
        cell.border    = source.border.copy()
        cell.fill      = source.fill.copy()
      if fill_formulae and source.data_type == Cell.TYPE_FORMULA:
        if source.value == "=":
          if source.coordinate in self.formula_attributes:
            fa = self.formula_attributes[source.coordinate].copy()
            self.formula_attributes[cell.coordinate] = fa
        else:
          # print("Copying formula from cell %s%d to %s%d"%(col,row-1,col,row))
          cell.value = re.sub(
            "(\$?[A-Z]{1,3})\$?%d"%(row - 1),
            lambda m: m.group(1) + str(row),
            source.value
          )   
        cell.data_type = Cell.TYPE_FORMULA
 
  # Check for Merged Cell Ranges that need to be expanded to contain new cells
  for cr_idx, cr in enumerate(self.merged_cell_ranges):
    g = RE_RANGE.search(cr).groupdict()
    #print("s_row = %s, e_row = %s, row = %d"%(g['s_row'], g['e_row'],row))
    if row_idx >= int(g['s_row']) and row_idx <= int(g['e_row']):
      #print("Expanding merged cell range '%s' by %d row(s) to '%d'"%(cr,cnt,int(g['e_row'])+cnt))
      self.merged_cell_ranges[cr_idx] = g['s_col'] + g['s_row'] + ":" + g['e_col'] + str(int(g['e_row'])+cnt)
 
  # Check for Formula Attributes that need reference ranges expanded to include new rows
  for k,v in self.formula_attributes.items():
    if 'ref' in v:
      ref = v['ref']
      if ":" in ref:
        g = RE_RANGE.search(v['ref']).groupdict()
        if row_idx >= int(g['s_row']) and row_idx <= int(g['e_row']):
          # print("Expanding cell range '%s' in formula attribute by %d row(s) to '%d'"%(ref,cnt,int(g['e_row'])+cnt))
          self.formula_attributes[k]['ref'] = g['s_col'] + g['s_row'] + ":" + g['e_col'] + str(int(g['e_row'])+cnt)

Worksheet.insert_rows = insert_rows

wb = load_workbook(filename = 'testInsert.xlsx')
ws = wb.worksheets[0]

ws.insert_rows(2,1, above=True) #below header cells single column of data

wb.save("testInsertColumnDone.xlsx")
