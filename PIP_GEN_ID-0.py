#libraries
import sys; print(sys.version) #build system
from datetime import datetime, date; print(datetime.now()) #date stamp
import xml.sax #.xml parser
import msvcrt #key press recognition
import openpyxl #create a workbook
from openpyxl.workbook import Workbook #inserting line in workbook
from openpyxl import load_workbook
from openpyxl.worksheet import *
from openpyxl.cell import  Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors







mcu = ['PSOC', 'NRF51', 'RENESAS', 'XILOG'];

equipment = ['TREADMILL_DEVICE', 'INCLINE_TRAINER_DEVICE', 'ELLIPTICAL_DEVICE',
             'EXERCISE_BIKE_DEVICE', 'SPIN_BIKE_DEVICE', 'VERTICAL_ELLIPTICAL_DEVICE',
             'FITNESS_WT_MACH_DEVICE', 'FREE_STRIDER_DEVICE', 'ROWER_DEVICE'];
          
display = ['NO_DISPLAY', 'SISP_DISPLAY', 'AW_DISPLAY', 'AW_DESK', 'RICKENBACKER_DISPLAY',
           'RICKENBACKER_MINI_DISPLAY', 'PM210_DISPLAY_PAUL_REED_SMITH', 'PM210_DISPLAY_GIBSON_NORDIC',
           'POWER_RING_DISPLAY', 'POWER_RING_DISPLAY_LARGE', 'AERO_POWER_RING_DISPLAY','TTW_DISPLAY',
           'LED_900', 'CD_DISPLAY', 'FENDER_DISPLAY', 'OFENDER_DISPLAY', 'OP_DISPLAY', 'IF17_DISPLAY',
           'IF17_VERT_DISPLAY'];

controller = ['TREAD_TACH', 'MC1650LS_2W', 'MC1618DLS', 'MC1648DLS', 'MC1618DHB', 'MC1648DHB', 'MC1618IHS',#7
              'MC1648IHS', 'MC1618IHB', 'MC1648IHB', 'MC2100LT_12', 'MC2100LTS_30', 'MC2100LTS_50W', 'MC2100ELS_18W_2Y',#14
              'MC2100ELS_50W_2Y', 'MC5100DTS_18W', 'MC5100DTS_50W', 'MC5100DTS3_50W', 'MC5100EDS_50W_V1',#19
              'MC5100EDS3_50W_V1', 'MC5150HCL', 'OLYMPUS_V1', 'RHYMBUS_DOMESTIC', 'RHYMEBUS_EUROPEAN', 'BASIC_BIKE',#25
              'PB_INC_18W', 'PB_INC_50W', 'PB_INC_485_18W', 'PB_INC_485_50W', 'PBCLA_2X50W_75W', 'PBCLA_FM_32_12V',#31
              'PB_PPI_485_18W', 'PB_PPI_485_48W']; #Rhymebus not used in this

voltage = ['0', '6', '9', '12'];
resistance = ['ADC'];
distance = ['LoResTach'];

maintainance = ['SOFTWARE_VERSION', 'CALIBRATE', 'FILL_FASTER', 'KEYCODE_TEST']; #tablet consoles will be different
         
ble = ['NO_PROFILE_SUPPORTED', 'WEIGHT_MACHINE_SUPPORTED', 'SPEED_RING_SUPPORTED', 'FILE_SYSTEM_SUPPORTED',
       'FITPRO_SYSTEM_SUPPORTED', 'HEART_RATE_SENSOR_SUPPORTED'];

pulse = ['Hand', 'Thumb', 'Chest', 'Ant', 'BLE', 'nanoHand', 'Priority'];
      
fan = ['TwoWire', 'ThreeWire', 'twoAndThree'];

audio = ['PC', 'BrainBoard', 'MP3', 'iPod', 'TV', 'FM', 'Headphones']; #may need upper or lower case adjustments

header = ['void for function call', 'RecordID', 'Link to Parts - Consoles', 'Link to Parts - Console Pics',
          'Part Number','Description', 'ProductType', 'Display Type', 'Equipment Needed', 'Setup', 'BLE SETUP',
          'Part Type Pull', 'PIP', 'Video of Inspection', 'EQF1259 Config', 'BLE Setup Image 1', 'BLE Setup Image 2',
          'Setup Image 1', 'Setup Image 2', 'ESD Bag', 'Cosmetics', 'Warning Label', 'Warning Label Picture',
          'Cosmetic, Color, and Texture', 'Picture of Console', 'Soldering Quality', 'Power up', 'Safety Key',
          'BLE Test', 'BLE Test Image 1', 'BLE Test Image 2', 'Software Version', 'Test Mode Image',
          'Display Test', 'Display Test Image', 'Button Test', 'Button TestImage', 'Drive Motor Output Test',
          'Tach Input Test', 'Resistance Motor Test', 'Incline Motor Test', 'USB port test', 'Hand Pulse Test',
          'Chest Pulse Test', 'Fan Test', 'Audio Test', 'TV Test', 'Upright Motor Test', 'Finish Test',
          'Last User', 'Created By User', 'Updated', 'Created'];



#EQF1259 source and to
app = load_workbook(filename = 'EQF1259 App Procedures.xlsx')
print(app.get_sheet_names()) #debug




#Read sheets
mat = app.get_sheet_by_name('matrix') #PIP database
eng = app.get_sheet_by_name('english')
chi = app.get_sheet_by_name('chinese')


#Create standalone
mywb = openpyxl.Workbook()
mywb.remove_sheet(mywb.get_sheet_by_name('Sheet')) #removes default sheet title
mywb.create_sheet(index=0, title='Parts - Consoles.csv') #adds new sheet with title
mywb.create_sheet(index=1, title='revision')


sta = mywb.get_sheet_by_name('Parts - Consoles.csv') #pip
for i in range(1, 53):
    _ = sta.cell(column = i, row = 1, value = header[i]) #write headers

for col in sta.columns: 
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = 100
     sta.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string


rev = mywb.get_sheet_by_name('revision') #rev
#Preliminary 
rev['A1'] = 'REV 0'
rev['B1'] = '20171005'
rev['C1'] = 'From chaos, springs new beginnings!'
rev['D1'] = 'ID_0'

rev['A2'] = 'REV 1'
rev['B2'] = '20171205'
rev['C2'] = 'Created arrays to add functionality for individual variables'
rev['D2'] = 'ID_0'

rev['A3'] = 'REV 2'
rev['B3'] = '20171206'
rev['C3'] = 'Added time stamping and found no use but will leave in'
rev['D3'] = 'ID_0'

rev['A4'] = 'REV 3'
rev['B4'] = '20171207'
rev['C4'] = 'Functions on test x.xml'
rev['D4'] = 'ID_0'

rev['A5'] = 'REV 4'
rev['B5'] = '20171227'
rev['C5'] = 'Converted to openpyxl and functions on test x.xml'
rev['D5'] = 'ID_0'

rev['A6'] = 'REV'
rev['B6'] = '20180112'
rev['C6'] = 'Killed trackvia full table and will upload individual as created'
rev['D6'] = 'ID_0'

rev['A7'] = 'REV'
rev['B7'] = datetime.now()
rev['C7'] = 'Let justice be done, though the heavens fall!'
rev['D7'] = 'ID_0'



for col in rev.columns: 
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2 #add defined width value
     rev.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string

print(mywb.get_sheet_names()) #debug
print rev['a5'].value, rev['c5'].value



           

#Parse .xml file
class Device( xml.sax.ContentHandler ):
   def __init__(self):
      
      self.CurrentData = ""
      self.consolePartNumber = ""
      self.mcuChipName = ""
      self.equipmentType = ""
      self.systemUnitType = ""
      self.BuildModelString = ""
      self.TypeName = ""
      self.PowerBroad = ""
      self.ConsoleVoltage = ""
      self.GradeProtocol = ""
      self.TabletProtocol = ""
      self.MaintenanceConfigFunction = ""
      self.PulseDriverItem = ""
      self.FanProtocol = ""
      self.AudioSrcItem = ""
      self.DistanceDriver = ""
      self.ResistanceDriver = ""
      
      



   # Call when an element starts
   def startElement(self, tag, attributes):
      self.CurrentData = tag
      if tag == "Config":
         print "---------System_Info---------\n"
      elif tag == "FeatureItem": #leave this in for debugging
         feature = attributes["xsi:type"]
         #print feature
         if feature == "UsbFeature":
             e = eng['s2'].value #USB
             c = chi['s2'].value
             sta['ao2'].alignment = Alignment(wrap_text = True)
             sta['ao2'] = e
                     
    # Call when a character is read
   def characters(self, content):
      if self.CurrentData == "consolePartNumber":
         self.consolePartNumber = content
      elif self.CurrentData == "BuildModelString":
         self.BuildModelString = content
      elif self.CurrentData == "equipmentType":
         self.equipmentType = content
      elif self.CurrentData == "TypeName":
         self.TypeName = content
      elif self.CurrentData == "mcuChipName":
         self.mcuChipName = content
      elif self.CurrentData == "systemUnitType":
         self.systemUnitType = content
      elif self.CurrentData == "PowerBoard":
         self.PowerBoard = content
      elif self.CurrentData == "ConsoleVoltage":
         self.ConsoleVoltage = content
      elif self.CurrentData == "GradeProtocol":
         self.GradeFeature = content
      elif self.CurrentData == "ResistanceDriver":
         self.ResistanceDriver = content
      elif self.CurrentData == "DistanceDriver":
         self.DistanceDriver = content
      elif self.CurrentData == "TabletProtocol":
         self.TabletProtocol = content
      elif self.CurrentData == "MaintenanceConfigFunction":
         self.MaintenanceConfigFunction = content
      elif self.CurrentData == "PulseDriverItem":
         self.PulseDriverItem = content
      elif self.CurrentData == "FanProtocol":
         self.FanProtocol = content
      elif self.CurrentData == "AudioSrcItem":
         self.AudioSrcItem = content
      
      

   # Call when an elements ends
   def endElement(self, tag):
      if self.CurrentData == "consolePartNumber":
         print "Part Number:", self.consolePartNumber
         sta['d2'].alignment = Alignment(wrap_text = True)
         sta['d2'] = self.consolePartNumber
      elif self.CurrentData == "BuildModelString":
         print "Part Name:", self.BuildModelString
         sta['e2'] = self.BuildModelString
      elif self.CurrentData == "equipmentType":
            print "Equipment:", self.equipmentType
            sta['f2'] = self.equipmentType
            i = 0
            for i in range(2):
               if self.equipmentType == equipment[i]:
                  e = eng['f2'].value #Warning label
                  c = chi['f2'].value
                  sta['u2'].alignment = Alignment(wrap_text = True)
                  sta['u2'] = e
                  e = eng['i2'].value #DMK
                  c = chi['i2'].value
                  sta['aa2'].alignment = Alignment(wrap_text = True)
                  sta['aa2'] = e
                  e = eng['o2'].value #Drive motor (PWM)
                  c = chi['o2'].value
                  sta['ak2'].alignment = Alignment(wrap_text = True)
                  sta['ak2'] = e
                  
      elif self.CurrentData == "TypeName": #Isolates tag to array items only
          if self.TypeName == display[0]:
              print "Display:", self.TypeName
              sta['g2'] = self.TypeName
          for i in range(1, 19):
              if self.TypeName == display[i]:
                  print "Display:", self.TypeName
                  sta['g2'] = self.TypeName
                  e = eng['g2'].value #Cosmetic all other product
                  c = chi['g2'].value
                  sta['w2'].alignment = Alignment(wrap_text = True)
                  sta['w2'] = e
      elif self.CurrentData == "mcuChipName":
         print "Processor:", self.mcuChipName
         if self.mcuChipName == mcu[1]:
            e = eng['j2'].value #BLE test
            c = chi['j2'].value
            sta['ab2'].alignment = Alignment(wrap_text = True)
            sta['ab2'] = e
            e = eng['t5'].value #BLE Chest pulse
            c = chi['t5'].value
            sta['aq2'].alignment = Alignment(wrap_text = True)
            sta['aq2'] = e
            print e
      elif self.CurrentData == "systemUnitType":
         print "Unit Measure:", self.systemUnitType
      elif self.CurrentData == "PowerBoard":
         print "Controller:", self.PowerBoard, '\n'
         #i = 0
         for i in range(2):
            if self.PowerBoard == controller[i]:
                e = eng['c3'].value #Tread 3
                c = chi['c3'].value
                sta['i2'].alignment = Alignment(wrap_text = True)
                sta['i2'] = e
         #i = 2
         for i in range(2, 20):
            if self.PowerBoard == controller[i]:
                ef = eng['c2'].value #Tread 1, Tread 2 will require verification of TV for Upright feature
                c = chi['c2'].value
                ep = eng['h2'].value #Tread 1, Tread 2 will require verification of TV for Upright feature
                c = chi['h2'].value
                sta['i2'].alignment = Alignment(wrap_text = True)
                sta['i2'] = ef
                sta['z2'].alignment = Alignment(wrap_text = True)
                sta['z2'] = ep
         if self.PowerBoard == controller[20]:
            e = eng['c6'].value #Boston 1, Boston 2 will require verification of TV or Upright feature
            c = chi['c6'].value
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = e
         if self.PowerBoard == controller[21]:
            e = eng['c5'].value #Club
            c = chi['c5'].value
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = e
         """for i in range(25, 26):
            if self.PowerBoard == controller[i]:
                e = eng['C6'].value #PB_INC
                c = chi['C6'].value
                sta['i2'].alignment = Alignment(wrap_text = True)
                sta['I2'] = e
               note = xlwt.easyxf('align: wrap yes')
               e = english.cell(0,0).value #PB_INC
               c = chinese.cell(0,0).value
               sheet1.write(1,8,(c+'\n'+e), note)
         i = 25
         for i in range(25):
            if self.PowerBoard == controller[i]:
               note = xlwt.easyxf('align: wrap yes')
               e = english.cell(0,0).value #PB_INC
               c = chinese.cell(0,0).value
               sheet1.write(1,8,(c+'\n'+e), note)
         i = 26
         for i in range(26):
            if self.PowerBoard == controller[i]:
               note = xlwt.easyxf('align: wrap yes')
               e = english.cell(0,0).value #PB_INC_485
               c = chinese.cell(0,0).value
               sheet1.write(1,8,(c+'\n'+e), note)"""
      elif self.CurrentData == "TabletProtocol":
          print "", self.TabletProtocol
          e = eng['g3'].value #Wolf cosmetic
          c = chi['g3'].value
          sta['w2'].alignment = Alignment(wrap_text = True)
          sta['w2'] = e
          e = eng['t6'].value #Wolf chest pulse
          c = chi['t6'].value
          sta['aq2'].alignment = Alignment(wrap_text = True)
          sta['aq2'] = e
          e = eng['w2'].value #Wolf HDMI
          c = chi['w2'].value
          sta['at2'].alignment = Alignment(wrap_text = True)
          sta['at2'] = e
      elif self.CurrentData == "GradeProtocol":
         print "", self.GradeProtocol
         e = eng['q2'].value #Incline
         c = chi['q2'].value
         sta['an2'].alignment = Alignment(wrap_text = True)
         sta['an2'] = e
      elif self.CurrentData == "MaintenanceConfigFunction":
         print "", self.MaintenanceConfigFunction
      elif self.CurrentData == "PulseDriverItem":
         print "", self.PulseDriverItem
         if self.PulseDriverItem == pulse[5] or pulse[0]:
             e = eng['t2'].value #Hand or nanoHand pulse
             c = chi['t2'].value
             sta['ap2'].alignment = Alignment(wrap_text = True)
             sta['ap2'] = e
             print e
         if self.PulseDriverItem == pulse[1]:
             e = eng['t3'].value #Thumb pulse
             c = chi['t3'].value
             sta['ap2'].alignment = Alignment(wrap_text = True)
             sta['ap2'] = e
             print e
         if self.PulseDriverItem == pulse[2]:
             e = eng['t4'].value #Chest pulse
             c = chi['t4'].value
             sta['aq2'].alignment = Alignment(wrap_text = True)
             sta['aq2'] = e
             print e
         if self.PulseDriverItem == pulse[3]:
             e = eng['t4'].value #ANT+ pulse
             c = chi['t4'].value
             sta['aq2'].alignment = Alignment(wrap_text = True)
             sta['aq2'] = e
             print e
      elif self.CurrentData == "FanProtocol":
         print "", self.FanProtocol
         if self.FanProtocol == fan[0]:
            e = eng['u2'].value #Two pin
            c = chi['u2'].value
            sta['ar2'].alignment = Alignment(wrap_text = True)
            sta['ar2'] = e
         else:
            e = eng['u3'].value #Three pin
            c = chi['u3'].value
            sta['ar2'].alignment = Alignment(wrap_text = True)
            sta['ar2'] = e
      elif self.CurrentData == "AudioSrcItem":
         print "", self.AudioSrcItem
         if self.AudioSrcItem == audio[1]:
             e = eng['v2'].value #Connect iPOD
             c = chi['v2'].value
             sta['as2'].alignment = Alignment(wrap_text = True)
             sta['as2'] = e
         if self.AudioSrcItem == audio[6]:
             e = eng['v3'].value #Connect headphones
             c = chi['v3'].value
             sta['as2'].alignment = Alignment(wrap_text = True)
             sta['as2'] = e
         if self.AudioSrcItem == audio[4]:
             e = eng['t3'].value #Connect TV
             c = chi['t3'].value
             sta['as2'].alignment = Alignment(wrap_text = True)
             sta['as2'] = e
      elif self.CurrentData == "ResistanceDriver":
         print "", self.ResistanceDriver
         e = eng['r2'].value #resistance
         c = chi['r2'].value
         sta['am2'].alignment = Alignment(wrap_text = True)
         sta['am2'] = e
      elif self.CurrentData == "DistanceDriver":
         print "", self.DistanceDriver
      elif self.CurrentData == "ConsoleVoltage":
         print "", self.ConsoleVoltage
         if self.ConsoleVoltage != voltage[0]:
            e = eng['c9'].value #Basic Bike will be defined with 0V for BIKE0 resist and BIKE2 resist and tach
            c = chi['c9'].value
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = e
         if self.ConsoleVoltage == voltage[0]:
            e = eng['c8'].value #Basic Bike will be defined with 0V for BIKE0 resist and BIKE2 resist and tach
            c = chi['c8'].value
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = e
      self.CurrentData = ""
      



if ( __name__ == "__main__"):

   # enter the .xml file name including ".xml"
   filename = raw_input("Enter File Name:\n")

   # create an XMLReader
   parser = xml.sax.make_parser()

   # turn off namepsaces
   parser.setFeature(xml.sax.handler.feature_namespaces, 0)




   # override the default ContextHandler
   Handler = Device()
   parser.setContentHandler( Handler )

   parser.parse(filename)

e = eng['e2'].value #ESD bag
c = chi['e2'].value
sta['s2'].alignment = Alignment(wrap_text = True)
sta['s2'] = e

e = eng['aa2'].value #finish
c = chi['aa2'].value
sta['av2'].alignment = Alignment(wrap_text = True)
sta['av2'] = e

sta['ax2'] = 'PIP Generator'

mywb.save('Parts - Consoles.csv.xlsx') #save created workbook
print '---------Procedure is building...Done!---------\n'



#execfile("PIP_GEN_ID_0.py") #uncomment launch for repeat .xml verification
#execfile("C:\Python27\MyScripts\PIP_GEN_ID_0.py") #used to launch from python shell
