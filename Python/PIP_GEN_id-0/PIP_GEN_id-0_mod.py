#libraries that are used in this script
import sys; #print(sys.version) #build system
import os
from datetime import datetime, date; #print(datetime.now()) #date stamp
import xml.sax #.xml parser
import msvcrt #key press recognition
import openpyxl #create a workbook
from openpyxl.workbook import Workbook #inserting line in workbook
from openpyxl import load_workbook
from openpyxl.worksheet import *
from openpyxl.cell import  Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors




mcu = ['PSOC', 'NRF51', 'RENESAS', 'XILOG'];

equipment = ['TREADMILL_DEVICE', 'INCLINE_TRAINER_DEVICE', 'ELLIPTICAL_DEVICE',
             'EXERCISE_BIKE_DEVICE', 'SPIN_BIKE_DEVICE', 'VERTICAL_ELLIPTICAL_DEVICE',
             'FITNESS_WT_MACH_DEVICE', 'FREE_STRIDER_DEVICE', 'ROWER_DEVICE'];

display = ['NO_DISPLAY', 'SISP_DISPLAY', 'AW_DISPLAY', 'AW_DESK', 'RICKENBACKER_DISPLAY',
           'RICKENBACKER_MINI_DISPLAY', 'PM210_DISPLAY_PAUL_REED_SMITH', 'PM210_DISPLAY_GIBSON_NORDIC',
           'POWER_RING_DISPLAY', 'POWER_RING_DISPLAY_LARGE', 'AERO_POWER_RING_DISPLAY','TTW_DISPLAY',
           'LED_900', 'CD_DISPLAY', 'FENDER_DISPLAY', 'OFENDER_DISPLAY', 'OP_DISPLAY', 'IF17_DISPLAY',
           'IF17_VERT_DISPLAY'];

tablet = ['RoyalWolf', 'Leagacy', 'Leagacy_FreeMotion']

controller = ['TREAD_TACH', 'MC1650LS_2W', 'MC1618DLS', 'MC1648DLS', 'MC1618DHB', 'MC1648DHB', 'MC1618IHS',#6
              'MC1648IHS', 'MC1618IHB', 'MC1648IHB', 'MC2100LT_12', 'MC2100LTS_30', 'MC2100LTS_50W', 'MC2100ELS_18W_2Y',#13
              'MC2100ELS_50W_2Y', 'MC5100DTS_18W', 'MC5100DTS_50W', 'MC5100DTS3_50W', 'MC5100EDS_50W_V1',#18
              'MC5100EDS3_50W_V1', 'MC5150HCL', 'OLYMPUS_V1', 'RHYMBUS_DOMESTIC', 'RHYMEBUS_EUROPEAN', 'BASIC_BIKE',#24
              'PB_INC_18W', 'PB_INC_50W', 'PB_INC_485_18W', 'PB_INC_485_50W', 'PB_PPI_485_18W', 'PB_PPI_485_48W',#30
              'PBCLA_2X50W_75W', 'PBCLA_FM_32_12V']; #Rhymebus not used in this

voltage = ['0', '6', '9', '12'];
resistance = ['ADC'];
distance = ['LoResTach'];

maintainance = ['CALIBRATE', 'TOGGLE_DISPLAY_ON_OFF', 'KEYCODE_TEST']; #tablet consoles will be different

keycode = ['SETTINGS', 'PRIORITY_DISPLAY', 'MANUAL']

ble = ['NO_PROFILE_SUPPORTED', 'WEIGHT_MACHINE_SUPPORTED', 'SPEED_RING_SUPPORTED', 'FILE_SYSTEM_SUPPORTED',
       'FITPRO_SYSTEM_SUPPORTED', 'HEART_RATE_SENSOR_SUPPORTED'];

pulse = ['Hand', 'Thumb', 'Chest', 'Ant', 'BLE', 'nanoHand', 'Priority'];

fan = ['TwoWire', 'ThreeWire', 'twoAndThree'];

audio = ['PC', 'BrainBoard', 'MP3', 'iPod', 'TV', 'FM', 'Headphones']; #may need upper or lower case adjustments

header = ['void for function call', 'RecordID', 'Link to Parts - Consoles', 'Link to Parts - Console Pics',
          'Part Number','Description', 'ProductType', 'Display Type', 'Equipment Needed', 'Setup', 'BLE SETUP',
          'Part Type Pull', 'PIP', 'Video of Inspection', 'EQF1259 Config', 'BLE Setup Image 1', 'BLE Setup Image 2',
          'Setup Image 1', 'Setup Image 2', 'ESD Bag', 'Cosmetics', 'Warning Label', 'Warning Label Picture',
          'Cosmetic, Color, and Texture', 'Picture of Console', 'Soldering Quality', 'Power up', 'Safety Key',
          'BLE Test', 'BLE Test Image 1', 'BLE Test Image 2', 'Software Version', 'Test Mode Image', 'Incline Calibration',
          'Display Test', 'Display Test Image', 'Button Test', 'Button Test Image', 'Drive Motor Output Test',
          'Tach Input Test', 'Resistance Motor Test', 'Incline Motor Test', 'USB port test', 'Hand Pulse Test',
          'Chest Pulse Test', 'Fan Test', 'Audio Test', 'TV Test', 'Upright Motor Test', 'Finish Test',
          'Last User', 'Created By User', 'Updated', 'Created'];



#EQF1259 source and to
app = load_workbook(filename = 'EQF1259 App Procedures.xlsx')
#print(app.get_sheet_names()) #debug




#Read sheets
mat = app.get_sheet_by_name('matrix') #PIP database
eng = app.get_sheet_by_name('english')
chi = app.get_sheet_by_name('chinese')


#Create standalone
mywb = openpyxl.Workbook()
mywb.remove_sheet(mywb.get_sheet_by_name('Sheet')) #removes default sheet title
mywb.create_sheet(index=0, title='Parts - Consoles.csv') #adds new sheet with title
mywb.create_sheet(index=1, title='revision')


sta = mywb.get_sheet_by_name('Parts - Consoles.csv') #pip
for i in range(1, 54):
    _ = sta.cell(column = i, row = 1, value = header[i]) #write headers

for col in sta.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:

         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = 100
     sta.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string


rev = mywb.get_sheet_by_name('revision') #rev
#Preliminary
rev['A1'] = 'REV 0'
rev['B1'] = '20171005'
rev['C1'] = 'From chaos, springs new beginnings!'
rev['D1'] = 'id-0'

rev['A2'] = 'REV 1'
rev['B2'] = '20171205'
rev['C2'] = 'Created arrays for variables'
rev['D2'] = 'id-0'

rev['A3'] = 'REV 2'
rev['B3'] = '20171206'
rev['C3'] = 'Added time stamping'
rev['D3'] = 'id-0'

rev['A4'] = 'REV 3'
rev['B4'] = '20171207'
rev['C4'] = 'Functions on test x.xml'
rev['D4'] = 'id-0'

rev['A5'] = 'REV 4'
rev['B5'] = '20171227'
rev['C5'] = 'Converted to openpyxl'
rev['D5'] = 'id-0'

rev['A6'] = 'REV 5'
rev['B6'] = '20180112'
rev['C6'] = 'Killed trackvia full table'
rev['D6'] = 'id-0'

rev['A7'] = 'REV 6'
rev['B7'] = '20180123'
rev['C7'] = 'Added chinese and auto launch for PIP'
rev['D7'] = 'id-0'

rev['A8'] = 'REV'
rev['B8'] = datetime.now()
rev['C8'] = 'Let justice be done, though the heavens fall!'
rev['D8'] = 'id-0'



for col in rev.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2 #add defined width value
     rev.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string

#print(mywb.get_sheet_names()) #debug
print rev['a7'].value, rev['c7'].value





#Parse .xml file
class Device( xml.sax.ContentHandler ):
   def __init__(self):

      self.CurrentData = ""
      self.consolePartNumber = ""
      self.mcuChipName = ""
      self.equipmentType = ""
      self.systemUnitType = ""
      self.BuildModelString = ""
      self.TypeName = ""
      self.PowerBroad = ""
      self.ConsoleVoltage = ""
      self.GradeProtocol = ""
      self.TabletProtocol = ""
      self.MaintenanceConfigFunction = ""
      self.KeyCodeName = ""
      self.PulseDriverItem = ""
      self.FanProtocol = ""
      self.AudioSrcItem = ""
      self.DistanceDriver = ""
      self.ResistanceDriver = ""





   # Call when an element starts
   def startElement(self, tag, attributes):
      self.CurrentData = tag
      if tag == "Config":
         print "---------System_Info---------\n"
      elif tag == "FeatureItem": #leave this in for debugging
         feature = attributes["xsi:type"]
         #print feature
         if feature == "UsbFeature":
             e = eng['s2'].value #USB
             c = chi['s2'].value
             n = '\n\n'
             com = c + n + e
             sta['ap2'].alignment = Alignment(wrap_text = True)
             sta['ap2'] = com

    # Call when a character is read

   def characters(self, content):
      if self.CurrentData == "consolePartNumber":
         self.consolePartNumber = content
      elif self.CurrentData == "BuildModelString":
         self.BuildModelString = content
      elif self.CurrentData == "equipmentType":
         self.equipmentType = content
      elif self.CurrentData == "TypeName":
         self.TypeName = content
      elif self.CurrentData == "mcuChipName":
         self.mcuChipName = content
      elif self.CurrentData == "systemUnitType":
         self.systemUnitType = content
      elif self.CurrentData == "PowerBoard":
         self.PowerBoard = content
      elif self.CurrentData == "ConsoleVoltage":
         self.ConsoleVoltage = content
      elif self.CurrentData == "GradeProtocol":
         self.GradeProtocol = content
      elif self.CurrentData == "ResistanceDriver":
         self.ResistanceDriver = content
      elif self.CurrentData == "DistanceDriver":
         self.DistanceDriver = content
      elif self.CurrentData == "TabletProtocol":
         self.TabletProtocol = content
      elif self.CurrentData == "MaintenanceConfigFunction":
         self.MaintenanceConfigFunction = content
      elif self.CurrentData == "KeyCodeName":
         self.KeyCodeName = content
      elif self.CurrentData == "PulseDriverItem":
         self.PulseDriverItem = content
      elif self.CurrentData == "FanProtocol":
         self.FanProtocol = content
      elif self.CurrentData == "AudioSrcItem":
         self.AudioSrcItem = content



   # Call when an elements ends
   def endElement(self, tag):
      if self.CurrentData == "consolePartNumber":
         print "Part Number:", self.consolePartNumber
         sta['d2'].alignment = Alignment(wrap_text = True)
         sta['d2'] = self.consolePartNumber
      elif self.CurrentData == "BuildModelString":
         print "Part Name:", self.BuildModelString
         sta['e2'] = self.BuildModelString
      elif self.CurrentData == "equipmentType":
         print "Equipment:", self.equipmentType
         sta['f2'] = self.equipmentType
         i = 0
         for i in range(2):
             if self.equipmentType == equipment[i]:
                 e = eng['f2'].value #Warning label
                 c = chi['f2'].value
                 n = '\n\n'
                 com = c + n + e
                 sta['u2'].alignment = Alignment(wrap_text = True)
                 sta['u2'] = com
                 e = eng['i2'].value #DMK
                 c = chi['i2'].value
                 com = c + n + e
                 sta['aa2'].alignment = Alignment(wrap_text = True)
                 sta['aa2'] = com
                 e = eng['o2'].value #Drive motor (PWM)
                 c = chi['o2'].value
                 com = c + n + e
                 sta['al2'].alignment = Alignment(wrap_text = True)
                 sta['al2'] = com      
         for i in range(2,8):
            if self.equipmentType == equipment[i]:
                 e = eng['p2'].value #Tach
                 c = chi['p2'].value
                 n = '\n\n'
                 com = c + n + e
                 sta['am2'].alignment = Alignment(wrap_text = True)
                 sta['am2'] = com
      elif self.CurrentData == "TypeName": #Isolates tag to array items only
         if self.TypeName == display[0]:
             print "Display:", self.TypeName
             sta['g2'] = self.TypeName
         for i in range(1, 19):
             if self.TypeName == display[i]:
                 print "Display:", self.TypeName
                 sta['g2'] = self.TypeName
                 e = eng['g2'].value #Cosmetic all other product
                 c = chi['g2'].value
                 n = '\n\n'
                 com = c + n + e
                 sta['w2'].alignment = Alignment(wrap_text = True)
                 sta['w2'] = com
      elif self.CurrentData == "mcuChipName":
         print "Processor:", self.mcuChipName
         if self.mcuChipName == mcu[1]:
             e = eng['j2'].value #BLE test
             c = chi['j2'].value
             n = '\n\n'
             com = c + n + e
             sta['ab2'].alignment = Alignment(wrap_text = True)
             sta['ab2'] = com
             e = eng['t5'].value #BLE pulse
             c = chi['t5'].value
             n = '\n\n'
             com = c + n + e
             sta['ar2'].alignment = Alignment(wrap_text = True)
             sta['ar2'] = com    
             if self.equipmentType == equipment[0] or equipment[1]:
                 e = eng['k5'].value #BLE maintenance
                 c = chi['k5'].value
                 n = '\n\n'
                 com = c + n + e
                 sta['ae2'].alignment = Alignment(wrap_text = True)
                 sta['ae2'] = com
                 print com
                 e = eng['l5'].value #BLE calibrate
                 c = chi['l5'].value
                 com = c + n + e
                 sta['ag2'].alignment = Alignment(wrap_text = True)
                 sta['ag2'] = com
                 e = eng['m5'].value #BLE display
                 c = chi['m5'].value
                 com = c + n + e
                 sta['ah2'].alignment = Alignment(wrap_text = True)
                 sta['ah2'] = com
                 e = eng['n5'].value #BLE keycodes
                 c = chi['n5'].value
                 com = c + n + e
                 sta['aj2'].alignment = Alignment(wrap_text = True)
                 sta['aj2'] = com
         if self.mcuChipName == mcu[2]:
             #if self.TabletProtocol == False:
                 if equipment[0] or equipment[1]:
                     e = eng['k6'].value #BLE maintenance
                     c = chi['k6'].value
                     n = '\n\n'
                     com = c + n + e
                     sta['ae2'].alignment = Alignment(wrap_text = True)
                     sta['ae2'] = com
                     print com
                     e = eng['l6'].value #BLE calibrate
                     c = chi['l6'].value
                     com = c + n + e
                     sta['ag2'].alignment = Alignment(wrap_text = True)
                     sta['ag2'] = com
                     e = eng['m6'].value #BLE display
                     c = chi['m6'].value
                     com = c + n + e
                     sta['ah2'].alignment = Alignment(wrap_text = True)
                     sta['ah2'] = com
                     e = eng['n6'].value #BLE keycodes
                     c = chi['n6'].value
                     com = c + n + e
                     sta['aj2'].alignment = Alignment(wrap_text = True)
                     sta['aj2'] = com

      elif self.CurrentData == "systemUnitType":
         print "Unit Measure:", self.systemUnitType
      elif self.CurrentData == "PowerBoard": #----------------------------------------------------------------------------------------
         print "Controller:", self.PowerBoard, '\n'
         for i in range(0,1):
            if self.PowerBoard == controller[i]:
                e = eng['c4'].value #Tread 3
                c = chi['c4'].value
                n = '\n\n'
                com = c + n + e
                sta['i2'].alignment = Alignment(wrap_text = True)
                sta['i2'] = com
         for i in range(2,20):
            if self.PowerBoard == controller[i]:
                ef = eng['c2'].value #Tread 1, Tread 2 will require verification of TV for Upright feature
                cf = chi['c2'].value
                n = '\n\n'
                comf = cf + n + ef
                ep = eng['h2'].value #Tread 1, Tread 2 will require verification of TV for Upright feature
                cp = chi['h2'].value
                n = '\n\n'
                comp = cp + n + ep
                sta['i2'].alignment = Alignment(wrap_text = True)
                sta['i2'] = comf
                sta['z2'].alignment = Alignment(wrap_text = True)
                sta['z2'] = comp
         if self.PowerBoard == controller[21]:
            ef = eng['c5'].value #Club treadmill
            cf = chi['c5'].value
            n = '\n\n'
            comf = cf + n + ef
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = comf
            ep = eng['h5'].value #Club treadmill
            cp = chi['h5'].value
            n = '\n\n'
            comp = cp + n + ep
            sta['z2'].alignment = Alignment(wrap_text = True)
            sta['z2'] = comp
         if self.PowerBoard == controller[20]:
            e = eng['c6'].value #Boston 1, Boston 2 will require verification of TV or Upright feature
            c = chi['c6'].value
            n = '\n\n'
            com = c + n + e
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = com
         if self.PowerBoard == controller[25] or self.PowerBoard == controller[26]:
            e = eng['c11'].value #PB_INC
            c = chi['c11'].value
            n = '\n\n'
            com = c + n + e
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = com
         for i in range(27,30):
            if self.PowerBoard == controller[i]:
                e = eng['c12'].value #PB_INC_485
                c = chi['c12'].value
                n = '\n\n'
                com = c + n + e
                sta['i2'].alignment = Alignment(wrap_text = True)
                sta['i2'] = com
         if self.PowerBoard == controller[31] or self.PowerBoard == controller[32]:
            e = eng['c11'].value #Club bike
            c = chi['c11'].value
            n = '\n\n'
            com = c + n + e
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = com
      elif self.KeyCodeName == "KeyCodeName":
          print "", self.KeyCodeName
      elif self.CurrentData == "TabletProtocol": #-------------------------------------------------------------
          print "", self.TabletProtocol
          if self.TabletProtocol == tablet[0]:
              e = eng['g3'].value #Wolf cosmetic
              c = chi['g3'].value
              n = '\n\n'
              com = c + n + e
              sta['w2'].alignment = Alignment(wrap_text = True)
              sta['w2'] = com
              e = eng['k2'].value #Wolf maintainance
              c = chi['k2'].value
              com = c + n + e
              sta['ae2'].alignment = Alignment(wrap_text = True)
              sta['ae2'] = com
              e = eng['l2'].value #Wolf calibrate
              c = chi['l2'].value
              com = c + n + e
              sta['ag2'].alignment = Alignment(wrap_text = True)
              sta['ag2'] = com
              e = eng['m2'].value #Wolf display
              c = chi['m2'].value
              com = c + n + e
              sta['ah2'].alignment = Alignment(wrap_text = True)
              sta['ah2'] = com
              #e = eng['n2'].value #Wolf keycodes not needed for this console style
              #c = chi['n2'].value
              #com = c + n + e
              #sta['aj2'].alignment = Alignment(wrap_text = True)
              #sta['aj2'] = com
              e = eng['t6'].value #Wolf chest pulse
              c = chi['t6'].value
              com = c + n + e
              sta['ar2'].alignment = Alignment(wrap_text = True)
              sta['ar2'] = com
              e = eng['w2'].value #Wolf HDMI
              c = chi['w2'].value
              com = c + n + e
              sta['au2'].alignment = Alignment(wrap_text = True)
              sta['au2'] = com
          if self.TabletProtocol == tablet[1]:#-----------------------------------------------------------
              e = eng['k3'].value #Legacy maintainance
              c = chi['k3'].value
              n = '\n\n'
              com = c + n + e
              sta['ae2'].alignment = Alignment(wrap_text = True)
              sta['ae2'] = com
              e = eng['l3'].value #Legacy calibrate
              c = chi['l3'].value
              com = c + n + e
              sta['ag2'].alignment = Alignment(wrap_text = True)
              sta['ag2'] = com
              e = eng['m2'].value #Legacy display
              c = chi['m2'].value
              com = c + n + e
              sta['ah2'].alignment = Alignment(wrap_text = True)
              sta['ah2'] = com
              e = eng['n3'].value #Legacy keycodes
              c = chi['n3'].value
              com = c + n + e
              sta['ai2'].alignment = Alignment(wrap_text = True)
              sta['ai2'] = com
              e = eng['t5'].value #Legacy Chest pulse
              c = chi['t5'].value
              n = '\n\n'
              com = c + n + e
              sta['ar2'].alignment = Alignment(wrap_text = True)
              sta['ar2'] = com
          if self.TabletProtocol == tablet[2]:#------------------------------------------------------------
              e = eng['k4'].value #Legacy FreeMotion maintainance
              c = chi['k4'].value
              n = '\n\n'
              com = c + n + e
              sta['ae2'].alignment = Alignment(wrap_text = True)
              sta['ae2'] = com
              e = eng['l4'].value #Legacy FreeMotion calibrate
              c = chi['l4'].value
              com = c + n + e
              sta['ag2'].alignment = Alignment(wrap_text = True)
              sta['ag2'] = com
              e = eng['m2'].value #Legacy FreeMotion display
              c = chi['m2'].value
              com = c + n + e
              sta['ah2'].alignment = Alignment(wrap_text = True)
              sta['ah2'] = com
              e = eng['n3'].value #Legacy FreeMotion keycodes
              c = chi['n3'].value
              com = c + n + e
              sta['ai2'].alignment = Alignment(wrap_text = True)
              sta['ai2'] = com
              e = eng['t4'].value #ANT pulse
              c = chi['t4'].value
              n = '\n\n'
              com = c + n + e
              sta['ar2'].alignment = Alignment(wrap_text = True)
              sta['ar2'] = com    
      elif self.CurrentData == "GradeProtocol":
         print "Incline:", self.GradeProtocol
         e = eng['q2'].value #Incline
         c = chi['q2'].value
         n = '\n\n'
         com = c + n + e
         sta['ao2'].alignment = Alignment(wrap_text = True)
         sta['ao2'] = com
      elif self.CurrentData == "MaintenanceConfigFunction":
         print "", self.MaintenanceConfigFunction
         
         """if self.MaintainanceConfigFunction == maintainance[0] and :
             e = eng['t2'].value #Calibrate
             c = chi['t2'].value
             n = '\n\n'
             com = c + n + e
             sta['aq2'].alignment = Alignment(wrap_text = True)
             sta['aq2'] = com"""






      elif self.CurrentData == "PulseDriverItem":
         print "", self.PulseDriverItem
         if self.PulseDriverItem == pulse[0] or self.PulseDriverItem == pulse[5]:
             e = eng['t2'].value #Hand or nanoHand pulse
             c = chi['t2'].value
             n = '\n\n'
             com = c + n + e
             sta['aq2'].alignment = Alignment(wrap_text = True)
             sta['aq2'] = com
         if self.PulseDriverItem == pulse[1]:
             e = eng['t3'].value #Thumb pulse
             c = chi['t3'].value
             n = '\n\n'
             com = c + n + e
             sta['aq2'].alignment = Alignment(wrap_text = True)
             sta['aq2'] = com    
         if self.PulseDriverItem == pulse[2]:
             e = eng['t4'].value #Chest pulse
             c = chi['t4'].value
             n = '\n\n'
             com = c + n + e
             sta['ar2'].alignment = Alignment(wrap_text = True)
             sta['ar2'] = com
      elif self.CurrentData == "FanProtocol":
         print "", self.FanProtocol
         if self.FanProtocol == fan[0]:
            e = eng['u2'].value #Two pin
            c = chi['u2'].value
            n = '\n\n'
            com = c + n + e
            sta['as2'].alignment = Alignment(wrap_text = True)
            sta['as2'] = com
         else:
            e = eng['u3'].value #Three pin
            c = chi['u3'].value
            n = '\n\n'
            com = c + n + e
            sta['as2'].alignment = Alignment(wrap_text = True)
            sta['as2'] = com
      elif self.CurrentData == "AudioSrcItem":
         print "", self.AudioSrcItem
         for i in range(0,4):
             if self.AudioSrcItem == audio[i]:
                e = eng['v2'].value #Connect iPOD
                c = chi['v2'].value
                n = '\n\n'
                com = c + n + e
                sta['at2'].alignment = Alignment(wrap_text = True)
                sta['at2'] = com
         if self.AudioSrcItem == audio[4]:
             e = eng['t3'].value #Connect TV
             c = chi['t3'].value
             n = '\n\n'
             com = c + n + e
             sta['at2'].alignment = Alignment(wrap_text = True)
             sta['at2'] = com
         if self.AudioSrcItem == audio[5]:
             e = eng['v4'].value #FM
             c = chi['v4'].value
             n = '\n\n'
             com = c + n + e
             sta['at2'].alignment = Alignment(wrap_text = True)
             sta['at2'] = com
         if self.AudioSrcItem == audio[6]:
             e = eng['v3'].value #Connect TV
             c = chi['v3'].value
             n = '\n\n'
             com = c + n + e
             sta['at2'].alignment = Alignment(wrap_text = True)
             sta['at2'] = com
      elif self.CurrentData == "ResistanceDriver":
         print "", self.ResistanceDriver
         e = eng['r2'].value #resistance
         c = chi['r2'].value
         n = '\n\n'
         com = c + n + e
         sta['an2'].alignment = Alignment(wrap_text = True)
         sta['an2'] = com
      elif self.CurrentData == "DistanceDriver":
         print "", self.DistanceDriver
      elif self.CurrentData == "ConsoleVoltage":
         print "", self.ConsoleVoltage
         if self.ConsoleVoltage != voltage[0]:
            ef = eng['c9'].value #Basic Bike1 with voltage > 0; resist and tach
            cf = chi['c9'].value
            n = '\n\n'
            comf = cf + n + ef
            ep = eng['h9'].value #Power up
            cp = chi['h9'].value
            n = '\n\n'
            comp = cp + n + ep
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = comf
            sta['z2'].alignment = Alignment(wrap_text = True)
            sta['z2'] = comp
         if self.ConsoleVoltage == voltage[0]:
            e = eng['c10'].value #Basic Bike2 voltage = 0; resist and tach
            c = chi['c10'].value
            n = '\n\n'
            com = c + n + e
            sta['i2'].alignment = Alignment(wrap_text = True)
            sta['i2'] = com
      self.CurrentData = ""
      


if ( __name__ == "__main__"):

   # enter the .xml file name including ".xml"
   filename = raw_input("Enter File Name:\n")

   # create an XMLReader
   parser = xml.sax.make_parser()

   # turn off namepsaces
   parser.setFeature(xml.sax.handler.feature_namespaces, 0)




   # override the default ContextHandler
   Handler = Device()
   parser.setContentHandler( Handler )

   parser.parse(filename)

e = eng['b2'].value #Equipment needed
c = chi['b2'].value
n = '\n\n'
com = c + n + e
sta['h2'].alignment = Alignment(wrap_text = True)
sta['h2'] = chi['b2'].value + '\n' + chi['b6'].value + '\n\n' + eng['b2'].value + '\n' + eng['b6'].value # Something new

e = eng['e2'].value #ESD bag
c = chi['e2'].value
n = '\n\n'
com = c + n + e
sta['s2'].alignment = Alignment(wrap_text = True)
sta['s2'] = com

e = eng['aa2'].value #finish
c = chi['aa2'].value
n = '\n\n'
com = c + n + e
sta['aw2'].alignment = Alignment(wrap_text = True)
sta['aw2'] = com

sta['ay2'] = 'PIP_GEN_id-0'
sta['ba2'] = datetime.now()




mywb.save('Parts - Consoles.csv.xlsx') #save created workbook
print '---------Procedure is building...Done!---------\n'




file = 'C:\\Users\\bryan.lee\\Documents\\GitHub\\hello-world\\Python\\PIP_GEN_id-0\\Parts - Consoles.csv.xlsx' #open procedure file
os.startfile(file)

execfile("PIP_GEN_id-0_mod.py") #uncomment launch for repeat .xml verification
#execfile('C:\Users\bryan.lee\Documents\GitHub\hello-world\Python\PIP_GEN_id-0\PIP_GEN_id-0_mod.py') #used to launch from python shell
#https://youtu.be/VKQ1Ph81Gps This is py2exe tutorial
