#libraries that are used in this script
import sys; print(sys.version, "\n\n")  #build system
print("-------------Welcome to PIP Generator-------------\n")
import os
from datetime import datetime, date; print(datetime.now()) #date stamp
import xml.sax #.xml parser
import openpyxl #create a workbook
from openpyxl.workbook import Workbook #inserting line in workbook
from openpyxl import load_workbook
from openpyxl.worksheet import *
from openpyxl.cell import  Cell
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors




mcu = ["PSOC", "NRF51", "RENESAS", "XILOG"];

equipment = ["TREADMILL_DEVICE", "INCLINE_TRAINER_DEVICE", "ELLIPTICAL_DEVICE",
             "EXERCISE_BIKE_DEVICE", "SPIN_BIKE_DEVICE", "VERTICAL_ELLIPTICAL_DEVICE",
             "FITNESS_WT_MACH_DEVICE", "FREE_STRIDER_DEVICE", "ROWER_DEVICE"];

display = ["NO_DISPLAY", "SISP_DISPLAY", "AW_DISPLAY", "AW_DESK", "RICKENBACKER_DISPLAY",
           "RICKENBACKER_MINI_DISPLAY", "PM210_DISPLAY_PAUL_REED_SMITH", "PM210_DISPLAY_GIBSON_NORDIC",
           "POWER_RING_DISPLAY", "POWER_RING_DISPLAY_LARGE", "AERO_POWER_RING_DISPLAY","TTW_DISPLAY",
           "LED_900", "CD_DISPLAY", "FENDER_DISPLAY", "OFENDER_DISPLAY", "OP_DISPLAY", "IF17_DISPLAY",
           "IF17_VERT_DISPLAY"];

tablet = ["RoyalWolf", "Legacy", "Legacy_FreeMotion"]

controller = ["TREAD_TACH", "MC1650LS_2W", "MC1618DLS", "MC1648DLS", "MC1618DHB", "MC1648DHB", "MC1618IHS",#6
              "MC1648IHS", "MC1618IHB", "MC1648IHB", "MC2100LT_12", "MC2100LTS_30", "MC2100LTS_50W", "MC2100ELS_18W_2Y",#13
              "MC2100ELS_50W_2Y", "MC5100DTS_18W", "MC5100DTS_50W", "MC5100DTS3_50W", "MC5100EDS_50W_V1",#18
              "MC5100EDS3_50W_V1", "MC5150HCL", "OLYMPUS_V1", "RHYMBUS_DOMESTIC", "RHYMEBUS_EUROPEAN", "BASIC_BIKE",#24
              "PB_INC_18W", "PB_INC_50W", "PB_INC_485_18W", "PB_INC_485_50W", "PB_PPI_485_18W", "PB_PPI_485_48W",#30
              "PBCLA_2X50W_75W", "PBCLA_FM_32_12V"]; #Rhymebus not used in this

voltage = ["0", "6", "9", "12"];
resistance = ["ADC"];
distance = ["LoResTach"];

maintainance = ["CALIBRATE", "TOGGLE_DISPLAY_ON_OFF", "KEYCODE_TEST"]; #tablet consoles will be different

keycode = ["SETTINGS", "PRIORITY_DISPLAY", "MANUAL"]

ble = ["NO_PROFILE_SUPPORTED", "WEIGHT_MACHINE_SUPPORTED", "SPEED_RING_SUPPORTED", "FILE_SYSTEM_SUPPORTED",
       "FITPRO_SYSTEM_SUPPORTED", "HEART_RATE_SENSOR_SUPPORTED"];

pulse = ["Hand", "Thumb", "Chest", "Ant", "BLE", "nanoHand", "Priority"];

fan = ["TwoWire", "ThreeWire", "twoAndThree"];

audio = ["None", "PC", "BrainBoard", "MP3", "iPod", "Headphones","BLE", "TV", "FM"]; #audio functions

header = ["void for function call", "RecordID", "Link to Parts - Console Pics", "Link to Parts - Consoles",
          "Part Number","Description", "ProductType", "Display Type", "Equipment Needed", "Setup", "BLE SETUP",
          "Part Type Pull", "PIP", "Video of Inspection", "EQF1259 Config", "BLE Setup Image 1", "BLE Setup Image 2",
          "Setup Image 1", "Setup Image 2", "ESD Bag", "Cosmetics", "Warning Label", "Warning Label Picture",
          "Cosmetic, Color, and Texture", "Picture of Console", "Soldering Quality", "Power up", "Safety Key",
          "BLE Test", "BLE Test Image 1", "BLE Test Image 2", "Software Version", "Test Mode Image", "Incline Calibration",
          "Display Test", "Display Test Image", "Button Test", "Button Test Image", "Drive Motor Output Test",
          "Tach Input Test", "Resistance Motor Test", "Incline Motor Test", "USB port test", "Hand Pulse Test",
          "Chest Pulse Test", "Fan Test", "Audio Test", "TV Test", "Upright Motor Test", "Finish Test",
          "Last User", "Created By User", "Updated", "Created"];



#EQF1259 source and to
app = load_workbook(filename = "ConsoleProcedures.xlsx")
#Read sheets
eng = app["english"] #PIP database
chi = app["chinese"]


#Create standalone
mywb = openpyxl.Workbook()
mywb.remove(mywb["Sheet"]) #removes default sheet title
mywb.create_sheet(index=0, title="Parts - Consoles") #adds new sheet with title
mywb.create_sheet(index=1, title="revision")

sta = mywb["Parts - Consoles"] #pip
for i in range(1, 54):
    _ = sta.cell(column = i, row = 1, value = header[i]) #write headers

for col in sta.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:

         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = 100
     sta.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string


rev = mywb["revision"] #rev
#Preliminary
rev["A1"] = "V.0"
rev["B1"] = "20171005"
rev["C1"] = "Order from chaos"
rev["D1"] = "id-0"

rev["A2"] = "V.1"
rev["B2"] = "20171205"
rev["C2"] = "Created arrays for variables"
rev["D2"] = "id-0"

rev["A3"] = "V.2"
rev["B3"] = "20171206"
rev["C3"] = "Added time stamping"
rev["D3"] = "id-0"

rev["A4"] = "V.3"
rev["B4"] = "20171207"
rev["C4"] = "Functions on test x.xml"
rev["D4"] = "id-0"

rev["A5"] = "V.4"
rev["B5"] = "20171227"
rev["C5"] = "Converted to openpyxl"
rev["D5"] = "id-0"

rev["A6"] = "V.5"
rev["B6"] = "20180112"
rev["C6"] = "Killed trackvia full table"
rev["D6"] = "id-0"

rev["A7"] = "V.6"
rev["B7"] = "20180123"
rev["C7"] = "Added chinese and auto launch for PIP"
rev["D7"] = "id-0"

rev["A8"] = "V.7"
rev["B8"] = "20180223"
rev["C8"] = "About done"
rev["D8"] = "id-0"

rev["A9"] = "V.8"
rev["B9"] = "20180228"
rev["C9"] = "Migrated to Python3.6.4"
rev["D9"] = "id-0"

rev["A10"] = "V.9"
rev["B10"] = "20180307"
rev["C10"] = "Test build with cx_Freeze5.1.1"
rev["D10"] = "id-0"

rev["A11"] = "V.x"
rev["B11"] = datetime.now()
rev["C11"] = "Fiat justitia ruat caelum "
rev["D11"] = "id-0"

print (rev["a10"].value, rev["c10"].value, "\n")


for col in rev.columns:
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2 #add defined width value
     rev.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string


#Parse .xml file
class Device( xml.sax.ContentHandler ):
   def __init__(self):

      self.CurrentData = ""
      self.consolePartNumber = ""
      self.mcuChipName = ""
      self.equipmentType = ""
      self.systemUnitType = ""
      self.BuildModelString = ""
      self.TypeName = ""
      self.PowerBoard = ""
      self.ConsoleVoltage = ""
      self.GradeProtocol = ""
      self.TabletProtocol = ""
      self.MaintenanceConfigFunction = ""
      self.KeyCodeName = ""
      self.PulseDriverItem = ""
      self.FanProtocol = ""
      self.AudioSrcItem = ""
      self.DistanceDriver = ""
      self.ResistanceDriver = ""
      self.KeyCodeCategory = ""



   # Call when an element starts
   def startElement(self, tag, attributes):
      self.CurrentData = tag
      if tag == "Config":
         print ("---------System_Info---------\n")
      elif tag == "FeatureItem": #feature tag reading
         feature = attributes["xsi:type"]
         #print feature
         if feature == "UsbFeature":
             print ("Feature:", feature)
             sta["ap2"].alignment = Alignment(wrap_text = True) #USB test
             sta["ap2"] = chi["s2"].value + "\n\n" + eng["s2"].value
             
         if feature == "CsafeFeature": #OT maintainance mode
             print ("Feature:", feature)
             sta["ae2"].alignment = Alignment(wrap_text = True) #OT maintainance
             sta["ae2"] = chi["k6"].value + "\n\n" + eng["k6"].value
             sta["ag2"].alignment = Alignment(wrap_text = True) #OT calibrate
             sta["ag2"] = chi["l6"].value + "\n\n" + eng["l6"].value
             sta["ah2"].alignment = Alignment(wrap_text = True) #OT display
             sta["ah2"] = chi["m6"].value + "\n\n" + eng["m6"].value
             sta["aj2"].alignment = Alignment(wrap_text = True) #OT keycodes
             sta["aj2"] = chi["n6"].value + "\n\n" + eng["n6"].value
         """if feature == "WahooFeature": #MYE CardioCare
             sta["???"].alignment = Alignment(wrap_text = True) #MYE CardioCare Add to TrackVia table options
             sta["???"] = chi["x2"].value + "\n\n" + eng["x2"].value
             sta["???"].alignment = Alignment(wrap_text = True) #OT BLE test Add to TrackVia table options
             sta["???"] = chi["y2"].value + "\n\n" + eng["y2"].value"""
         if feature == "TvFeature": #39818 TV test
             sta["au2"].alignment = Alignment(wrap_text = True) #MYE TV test
             sta["au2"] = chi["w4"].value + "\n\n" + eng["w4"].value
         if feature == "MaintenanceModeFeature":
             print ("Feature:", feature)
             if self.KeyCodeName == "SETTINGS":
              print ("something else---------------------------------------------------------------------")
             
         
    # Call when a character is read

   def characters(self, content):
      if self.CurrentData == "consolePartNumber":
         self.consolePartNumber = content
      elif self.CurrentData == "BuildModelString":
         self.BuildModelString = content
      elif self.CurrentData == "equipmentType":
         self.equipmentType = content
      elif self.CurrentData == "TypeName":
         self.TypeName = content
      elif self.CurrentData == "mcuChipName":
         self.mcuChipName = content
      elif self.CurrentData == "systemUnitType":
         self.systemUnitType = content
      elif self.CurrentData == "PowerBoard":
         self.PowerBoard = content
      elif self.CurrentData == "ConsoleVoltage":
         self.ConsoleVoltage = content
      elif self.CurrentData == "GradeProtocol":
         self.GradeProtocol = content
      elif self.CurrentData == "ResistanceDriver":
         self.ResistanceDriver = content
      elif self.CurrentData == "DistanceDriver":
         self.DistanceDriver = content
      elif self.CurrentData == "TabletProtocol":
         self.TabletProtocol = content
      elif self.CurrentData == "MaintenanceConfigFunction":
         self.MaintenanceConfigFunction = content
      elif self.CurrentData == "KeyCodeName":
         self.KeyCodeName = content
      elif self.CurrentData == "PulseDriverItem":
         self.PulseDriverItem = content
      elif self.CurrentData == "FanProtocol":
         self.FanProtocol = content
      elif self.CurrentData == "AudioSrcItem":
         self.AudioSrcItem = content
      elif self.CurrentData == "KeyCodeCategory":
         self.KeyCodeCategory = content



   # Call when an elements ends
   def endElement(self, tag):
      if self.CurrentData == "consolePartNumber":
         print ("Part Number:", self.consolePartNumber)
         sta["d2"].alignment = Alignment(wrap_text = True)
         sta["d2"] = self.consolePartNumber
      elif self.CurrentData == "BuildModelString":
         print ("Part Name:", self.BuildModelString)
         sta["e2"].alignment = Alignment(wrap_text = True)
         sta["e2"] = self.BuildModelString
      elif self.CurrentData == "equipmentType":
         print ("Equipment:", self.equipmentType)
         sta["f2"].alignment = Alignment(wrap_text = True)
         sta["f2"] = self.equipmentType
         i = 0
         for i in range(2):
             if self.equipmentType == equipment[i]:
                 sta["u2"].alignment = Alignment(wrap_text = True) #Warning label
                 sta["u2"] = chi["f2"].value + "\n\n" + eng["f2"].value
                 sta["aa2"].alignment = Alignment(wrap_text = True) #DMK
                 sta["aa2"] = chi["i2"].value + "\n\n" + eng["i2"].value
                 sta["al2"].alignment = Alignment(wrap_text = True) #Drivemotor
                 sta["al2"] = chi["o2"].value + "\n\n" + eng["o2"].value
         for i in range(2,8):
            if self.equipmentType == equipment[i]:
                 sta["am2"].alignment = Alignment(wrap_text = True) #Tach
                 sta["am2"] = chi["p2"].value + "\n\n" + eng["p2"].value
      elif self.CurrentData == "TypeName": #Isolates tag to array items only
         if self.TypeName == display[0]:
             print ("Display:", self.TypeName)
             sta["g2"].alignment = Alignment(wrap_text = True)
             sta["g2"] = self.TypeName
         for i in range(1, 19):
             if self.TypeName == display[i]:
                 print ("Display:", self.TypeName)
                 sta["g2"].alignment = Alignment(wrap_text = True)
                 sta["g2"] = self.TypeName
                 sta["w2"].alignment = Alignment(wrap_text = True) #Cosmetic all other product
                 sta["w2"] = chi["g2"].value + "\n\n" + eng["g2"].value
      elif self.CurrentData == "mcuChipName":
         print ("Processor:", self.mcuChipName)
         if self.mcuChipName == mcu[1]:
             sta["ab2"].alignment = Alignment(wrap_text = True) #BLE test
             sta["ab2"] = chi["j2"].value + "\n\n" + eng["j2"].value
             sta["ar2"].alignment = Alignment(wrap_text = True) #BLE pulse
             sta["ar2"] = chi["t5"].value + "\n\n" + eng["t5"].value
             if self.equipmentType == ("TREADMILL_DEVICE" or "INCLINE_TRAINER_DEVICE"):
                 sta["ae2"].alignment = Alignment(wrap_text = True) #BLE maintainance
                 sta["ae2"] = chi["k5"].value + "\n\n" + eng["k5"].value
                 sta["ag2"].alignment = Alignment(wrap_text = True) #BLE calibrate
                 sta["ag2"] = chi["l5"].value + "\n\n" + eng["l5"].value
                 sta["ah2"].alignment = Alignment(wrap_text = True) #BLE display
                 sta["ah2"] = chi["m5"].value + "\n\n" + eng["m5"].value
                 sta["aj2"].alignment = Alignment(wrap_text = True) #BLE keycodes
                 sta["aj2"] = chi["n5"].value + "\n\n" + eng["n5"].value
      elif self.CurrentData == "systemUnitType":
         print ("Unit Measure:", self.systemUnitType)
      elif self.CurrentData == "PowerBoard": #----------------------------------------------------------------------------------------
         print ("Controller:", self.PowerBoard, "\n")
         for i in range(0,1):
            if self.PowerBoard == controller[i]:
                sta["i2"].alignment = Alignment(wrap_text = True) #Tread 3
                sta["i2"] = chi["c4"].value + "\n\n" + eng["c4"].value
         for i in range(2,20):
            if self.PowerBoard == controller[i]:
                sta["i2"].alignment = Alignment(wrap_text = True) #Fixture Callout
                sta["i2"] = chi["c2"].value + "\n\n" + eng["c2"].value
                sta["z2"].alignment = Alignment(wrap_text = True) #Power up
                sta["z2"] = chi["h2"].value + "\n\n" + eng["h2"].value
                sta["h2"].alignment = Alignment(wrap_text = True) #Equipment needed
                sta["h2"] = chi["b2"].value + "\n" + chi["b7"].value + "\n\n" + eng["b2"].value + "\n" + eng["b7"].value
         if self.PowerBoard == controller[21]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Club fixture Callout
            sta["i2"] = chi["c5"].value + "\n\n" + eng["c5"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h5"].value + "\n\n" + eng["h5"].value
         if self.PowerBoard == controller[20]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Boston 1, Boston 2 will require verification of TV or Upright feature
            sta["i2"] = chi["c6"].value + "\n\n" + eng["c6"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h2"].value + "\n\n" + eng["h2"].value
         if self.PowerBoard == controller[25] or self.PowerBoard == controller[26]:
            sta["i2"].alignment = Alignment(wrap_text = True) #PB_INC
            sta["i2"] = chi["c11"].value + "\n\n" + eng["c11"].value
         for i in range(27,30):
            if self.PowerBoard == controller[i]:
                sta["i2"].alignment = Alignment(wrap_text = True) #PB_INC_485
                sta["i2"] = chi["c12"].value + "\n\n" + eng["c12"].value
         if self.PowerBoard == controller[31] or self.PowerBoard == controller[32]:
            sta["i2"].alignment = Alignment(wrap_text = True) #CLUB_BIKE
            sta["i2"] = chi["c11"].value + "\n\n" + eng["c11"].value
      elif self.CurrentData == "KeyCodeName":
          if self.KeyCodeName == "SETTINGS":
              print ("KeyName:", self.KeyCodeName)
          if self.KeyCodeName == "MANUAL":
              print ("KeyName:", self.KeyCodeName)
          if self.KeyCodeName == "PRIORITY_DISPLAY":
              print ("KeyName:", self.KeyCodeName)
      elif self.CurrentData == "KeyCodeCategory":
          if self.KeyCodeName == "SETTINGS":
              sta["ae2"].alignment = Alignment(wrap_text = True) #software
              sta["ae2"] = chi["k8"].value + "\n\n" + eng["k8"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #incline
              sta["ag2"] = chi["l8"].value + "\n\n" + eng["l8"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #display
              sta["ah2"] = chi["m8"].value + "\n\n" + eng["m8"].value
              sta["aj2"].alignment = Alignment(wrap_text = True) #keycodes
              sta["aj2"] = chi["n8"].value + "\n\n" + eng["n8"].value
          if self.GradeProtocol != "Manual":
              sta["ag2"].alignment = Alignment(wrap_text = True) #incline
              sta["ag2"] = ""
      elif self.CurrentData == "KeyCodeCategory":
          if (self.KeyCodeName == "MANUAL" or "PRIORITY_DISPLAY"):
              sta["ae2"].alignment = Alignment(wrap_text = True) #software
              sta["ae2"] = chi["k9"].value + "\n\n" + eng["k9"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #incline
              sta["ag2"] = chi["l9"].value + "\n\n" + eng["l9"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #display
              sta["ah2"] = chi["m9"].value + "\n\n" + eng["m9"].value
              sta["aj2"].alignment = Alignment(wrap_text = True) #keycodes
              sta["aj2"] = chi["n9"].value + "\n\n" + eng["n9"].value
          if self.GradeProtocol != "Manual":
              sta["ag2"].alignment = Alignment(wrap_text = True) #incline
              sta["ag2"] = ""
      elif self.CurrentData == "TabletProtocol": #-------------------------------------------------------------
          print ("Tablet Type:", self.TabletProtocol)
          if self.TabletProtocol == tablet[0]:
              sta["w2"].alignment = Alignment(wrap_text = True) #Wolf cosmetics
              sta["w2"] = chi["g3"].value + "\n\n" + eng["g3"].value
              sta["ae2"].alignment = Alignment(wrap_text = True) #Wolf maintainance
              sta["ae2"] = chi["k2"].value + "\n\n" + eng["k2"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #Wolf calibrate
              sta["ag2"] = chi["l2"].value + "\n\n" + eng["l2"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #Wolf display
              sta["ah2"] = chi["m2"].value + "\n\n" + eng["m2"].value
              sta["ar2"].alignment = Alignment(wrap_text = True) #Wolf chest pulse
              sta["ar2"] = chi["t6"].value + "\n\n" + eng["t6"].value
              sta["au2"].alignment = Alignment(wrap_text = True) #Wolf HDMI
              sta["au2"] = chi["w2"].value + "\n\n" + eng["w2"].value
          if self.TabletProtocol == tablet[2]:#------------------------------------------------------------
              sta["ae2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion maintainance
              sta["ae2"] = chi["k4"].value + "\n\n" + eng["k4"].value
              sta["ag2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion calibrate
              sta["ag2"] = chi["l4"].value + "\n\n" + eng["l4"].value
              sta["ah2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion display
              sta["ah2"] = chi["m2"].value + "\n\n" + eng["m2"].value
              sta["ai2"].alignment = Alignment(wrap_text = True) #Legacy FreeMotion maintainance
              sta["ai2"] = chi["n3"].value + "\n\n" + eng["n3"].value
              sta["ar2"].alignment = Alignment(wrap_text = True) #ANT pulse
              sta["ar2"] = chi["t4"].value + "\n\n" + eng["t4"].value
      elif self.CurrentData == "GradeProtocol":
         print ("Incline Type:", self.GradeProtocol)
         sta["ao2"].alignment = Alignment(wrap_text = True) #Incline
         sta["ao2"] = chi["q2"].value + "\n\n" + eng["q2"].value
      elif self.CurrentData == "MaintenanceConfigFunction":
         print ("Maintaince Screens:", self.MaintenanceConfigFunction)
      elif self.CurrentData == "PulseDriverItem": #-------------------------------------------------------------------
         print ("Pulse Type:", self.PulseDriverItem)
         if self.PulseDriverItem == pulse[0] or self.PulseDriverItem == pulse[5]:
             sta["aq2"].alignment = Alignment(wrap_text = True) #Hand or nanoHand pulse
             sta["aq2"] = chi["t2"].value + "\n\n" + eng["t2"].value
         if self.PulseDriverItem == pulse[1]:
             sta["aq2"].alignment = Alignment(wrap_text = True) #Thumb pulse
             sta["aq2"] = chi["t3"].value + "\n\n" + eng["t3"].value
         if self.PulseDriverItem == pulse[2]:
             sta["ar2"].alignment = Alignment(wrap_text = True) #Chest pulse
             sta["ar2"] = chi["t4"].value + "\n\n" + eng["t4"].value
      elif self.CurrentData == "FanProtocol":
         print ("Fan Pin:", self.FanProtocol)
         if self.FanProtocol == fan[0]:
            sta["as2"].alignment = Alignment(wrap_text = True) #Two pin
            sta["as2"] = chi["u2"].value + "\n\n" + eng["u2"].value
         else:
            sta["as2"].alignment = Alignment(wrap_text = True) #Three pin
            sta["as2"] = chi["u3"].value + "\n\n" + eng["u3"].value
      elif self.CurrentData == "AudioSrcItem": #--------------------------------------------------------------------
         print ("Audio Source:", self.AudioSrcItem)
         if self.AudioSrcItem == audio[2]:
             sta["at2"].alignment = Alignment(wrap_text = True) #Connect audio source
             sta["at2"] = chi["v2"].value + "\n\n" + eng["v2"].value
         if self.AudioSrcItem == audio[5]:
             sta["at2"].alignment = Alignment(wrap_text = True) #Connect audio source and headphones
             sta["at2"] = chi["v3"].value + "\n\n" + eng["v3"].value
         if self.AudioSrcItem == audio[6]:
             sta["at2"].alignment = Alignment(wrap_text = True) #Connect audio source and headphones and BLE audio
             sta["at2"] = chi["v4"].value + "\n\n" + eng["v4"].value
         if self.AudioSrcItem == audio[8]:
             sta["at2"].alignment = Alignment(wrap_text = True) #Connect MYE audio
             sta["at2"] = chi["v5"].value + "\n\n" + eng["v5"].value
      elif self.CurrentData == "ResistanceDriver": #-----------------------------------------------------------
         print ("Resistance Type:",self.ResistanceDriver)
         sta["an2"].alignment = Alignment(wrap_text = True) #Resistance
         sta["an2"] = chi["r2"].value + "\n\n" + eng["r2"].value
      elif self.CurrentData == "DistanceDriver":
         print ("Tach type:", self.DistanceDriver)
      elif self.CurrentData == "ConsoleVoltage":
         print ("Console Voltage:", self.ConsoleVoltage)#---------------------------------------------------------------
         if self.ConsoleVoltage != voltage[0]:
            sta["i2"].alignment = Alignment(wrap_text = True) #Basic Bike1 with voltage > 0; resist and tach
            sta["i2"] = chi["c9"].value + "\n\n" + eng["c9"].value
            sta["z2"].alignment = Alignment(wrap_text = True) #Power up
            sta["z2"] = chi["h9"].value + "\n\n" + eng["h9"].value
         if self.ConsoleVoltage == voltage[0]:#------------------------------------------------------------------------
            sta["i2"].alignment = Alignment(wrap_text = True) #Basic Bike2 voltage = 0; resist and tach
            sta["i2"] = chi["c10"].value + "\n\n" + eng["c10"].value
      self.CurrentData = ""
      


if ( __name__ == "__main__"):

   # enter the .xml file name including ".xml"
   filename = input("Enter File Name:\n")

   # create an XMLReader
   parser = xml.sax.make_parser()

   # turn off namepsaces
   parser.setFeature(xml.sax.handler.feature_namespaces, 0)




   # override the default ContextHandler
   Handler = Device()
   parser.setContentHandler( Handler )

   parser.parse(filename)


sta["c2"].alignment = Alignment(wrap_text = True) #[PartTypeLink]
sta["c2"] = "控制台 console"

sta["s2"].alignment = Alignment(wrap_text = True) #ESD bag
sta["s2"] = chi["e2"].value + "\n\n" + eng["e2"].value

sta["aw2"].alignment = Alignment(wrap_text = True) #Finish
sta["aw2"] = chi["aa2"].value + "\n\n" + eng["aa2"].value
sta["ay2"].alignment = Alignment(wrap_text = True) #Finish
sta["ay2"] = "PIP_Generator"
sta["ba2"].alignment = Alignment(wrap_text = True) #Finish
sta["ba2"] = datetime.now()




mywb.save("Parts - Consoles.xlsx") #save created workbook. Can designate path for saving.
print ("\n")
print ("-------------Procedure Build Is Complete!-------------\n")

file = "C:\\Users\\bryan.lee\\Documents\\GitHub\\hello-world\\python3.6\\PIP_GEN_id-0\\Parts - Consoles.xlsx" #open procedure file based on the path called out in previous lines.
os.startfile(file)

#input ("Press Enter To Close Window:")
