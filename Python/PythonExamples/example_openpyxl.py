#datacamp.com example sheet
#libraries
import sys
import openpyxl
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color, colors




#load .xlsx
wb = load_workbook('C:\Python27\MyScripts\Examples\Parts - Consoles.csv.xlsx')

#sheet info
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Parts - Consoles.csv')
print sheet.title
anotherSheet = wb.active
print anotherSheet

#cell info 1
print sheet['H8'].value

c = sheet['B4']
print c.row
print c.column
print c.coordinate

#cell info 2
print sheet.cell(row=6, column=9).value
for i in range(1, 5):
    print(i, sheet.cell(row=i, column=7).value)
    
#openpyxl.utils
print get_column_letter(1)
print column_index_from_string('A')

for cellObj in sheet['A1' : 'C3']:
    for cell in cellObj:
        print(cell.coordinate, cell.value)
    print('---END---')
    
print sheet.max_row
print sheet.max_column

#create and write to new file with openpyxl
mywb = openpyxl.Workbook() #creates a sheet
mywb.create_sheet() #adds a sheet
print(mywb.get_sheet_names())

mywb.remove_sheet(mywb.get_sheet_by_name('Sheet')) #removes auto created and added sheets
mywb.remove_sheet(mywb.get_sheet_by_name('Sheet1'))
mywb.create_sheet(index=0, title='sheet_1') #adds sheet with title
mywb.create_sheet(index=1, title='sheet_2')
ms = mywb.get_sheet_by_name('sheet_1') #variable for sheet

for i in range(0,10,1):
    ms.row_dimensions[i].height = 120 #adjust multi row dimension
    ms.column_dimensions['F'].width = 60 #adjust column dimenstion

ms['H8'] = 'Found the Money Pit' #write to cell
var = 10
varst = 'variable string'
ms['A1'] = var
ms['A2'] = varst
mywb.save('New_xlsx_File.xlsx') #save new workbook

print(mywb.get_sheet_names())
print ms['H8'].value
print ms['A1'].value
print ms['A2'].value

#another create with openpyxl examples
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = 'range_names'
for row in range(1, 40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title='Pi')

ws2['F5'] = 3.14
ws2['A1'] = datetime.datetime(2017, 12, 28) #will not fit as variable
ws2['B4'] = 'Will this length fit?\nI hope it does and is not too long!' #will fit if it is a string

for col in ws2.columns: 
     max_length = 0
     column = col[0].column # Get the column name
     for cell in col:
         try: # Necessary to avoid error on empty cells
             if len(str(cell.value)) > max_length:
                 max_length = len(cell.value)
         except:
             pass
     adjusted_width = (max_length + 2) * 1.2 #add defined width value
     ws2.column_dimensions[column].width = adjusted_width #adjust multi column width that contains string

#ws2['B4'].alignment = Alignment(wrap_text = True)
ws2['C3'].value = "Line 1\nLine 2\nLine 3"
x = 'line1'
y = '\n'
z = 'line2'
com = x + y + z
ws2['d2'].value = com

ws3 = wb.create_sheet(title='data')
for row in range(10, 20):
    for col in range(27, 54):
        _=ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col))) #write to certain cells

wb.save(filename = dest_filename)

print(ws3['AA10'].value)


"""style defaults
font = Font(name='Calibri',
                 size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')
fill = PatternFill(fill_type=None,
                 start_color='FFFFFFFF',
                 end_color='FF000000')
border = Border(left=Side(border_style=None,
                           color='FF000000'),
                 right=Side(border_style=None,
                            color='FF000000'),
                 top=Side(border_style=None,
                          color='FF000000'),
                 bottom=Side(border_style=None,
                             color='FF000000'),
                 diagonal=Side(border_style=None,
                               color='FF000000'),
                 diagonal_direction=0,
                 outline=Side(border_style=None,
                              color='FF000000'),
                 vertical=Side(border_style=None,
                               color='FF000000'),
                 horizontal=Side(border_style=None,
                                color='FF000000')
                )
alignment=Alignment(horizontal='general',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
number_format = 'General'
protection = Protection(locked=True,
                         hidden=False)
"""

